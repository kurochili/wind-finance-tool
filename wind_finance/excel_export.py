"""
wind_finance.excel_export
=========================
导出格式化 Excel 报告（封面 + 参数表 + 逐年现金流表 + 财务指标汇总）
"""

from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from .calculator import CalculationResult
from .models import WindFarmFinancialInputs

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E79")
NUM_FMT_USD = '#,##0'
NUM_FMT_USD2 = '#,##0.00'
NUM_FMT_PCT = '0.00%'
NUM_FMT_LCOE = '0.000000'


def _style_cell(ws, row, col, value, font=BODY_FONT, fill=None, fmt=None, align_h="right"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align_h, vertical="center")
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def _write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        _style_cell(ws, row, col_start + i, h, font=HEADER_FONT, fill=HEADER_FILL, align_h="center")


def _auto_width(ws, min_width=12, max_width=25):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _create_cover(wb: Workbook, inputs: WindFarmFinancialInputs):
    """封面页"""
    ws = wb.active
    ws.title = "封面"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.merge_cells("B3:F3")
    cell = ws.cell(row=3, column=2, value="风电项目经济性评估报告")
    cell.font = Font(name="Arial", bold=True, size=24, color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:F5")
    ws.cell(row=5, column=2, value=inputs.basic.project_name).font = Font(
        name="Arial", size=16, color="333333"
    )
    ws["B5"].alignment = Alignment(horizontal="center")

    info = [
        ("项目类型", "海上风电" if inputs.basic.project_type == "offshore" else "陆上风电"),
        ("国家/地区", inputs.basic.country),
        ("装机容量", f"{inputs.capacity_mw:.0f} MW"),
        ("满负荷小时数", f"{inputs.basic.full_load_hours} h"),
        ("运营期", f"{inputs.operational.operation_years} 年"),
        ("货币单位", "USD"),
    ]
    for i, (label, val) in enumerate(info, start=8):
        ws.cell(row=i, column=2, value=label).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=i, column=3, value=val).font = Font(name="Arial", size=11)

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30


def _create_params_sheet(wb: Workbook, inputs: WindFarmFinancialInputs):
    """参数表"""
    ws = wb.create_sheet("输入参数")
    ws.sheet_properties.tabColor = "2E75B6"

    row = 1
    _write_header_row(ws, row, ["参数分类", "参数名称", "值", "单位"])

    sections = [
        ("风场基本信息", [
            ("项目名称", inputs.basic.project_name, ""),
            ("项目类型", inputs.basic.project_type, ""),
            ("国家/地区", inputs.basic.country, ""),
            ("机组台数", inputs.basic.num_turbines, "台"),
            ("单机容量", inputs.basic.turbine_capacity_mw, "MW"),
            ("总装机容量", inputs.capacity_mw, "MW"),
            ("满负荷小时数", inputs.basic.full_load_hours, "h"),
            ("综合厂用电及线损率", inputs.basic.loss_rate, "%"),
            ("年上网电量", inputs.net_annual_generation_mwh, "MWh"),
            ("建设期", inputs.basic.construction_months, "月"),
        ]),
        ("投资造价", [
            ("单位千瓦静态投资", inputs.investment.resolve_unit_investment(), "USD/kW"),
            ("静态总投资", inputs.total_static_investment, "USD"),
            ("建设期利息", inputs.construction_interest, "USD"),
            ("动态总投资", inputs.total_dynamic_investment, "USD"),
            ("流动资金", inputs.working_capital, "USD"),
            ("项目总投资", inputs.total_investment, "USD"),
        ]),
        ("融资条件", [
            ("资本金比例", inputs.financing.equity_ratio, "%"),
            ("贷款利率", inputs.financing.long_term_loan_rate, "%"),
            ("贷款期限", inputs.financing.loan_term_years, "年"),
            ("资本金金额", inputs.equity_amount, "USD"),
            ("贷款金额", inputs.debt_amount, "USD"),
        ]),
        ("运营假设", [
            ("运营期", inputs.operational.operation_years, "年"),
            ("折旧年限", inputs.operational.depreciation_years, "年"),
            ("残值率", inputs.operational.residual_rate, "%"),
            ("质保期", inputs.operational.warranty.warranty_years, "年"),
            ("定员人数", inputs.operational.staff_count, "人"),
            ("保险费率", inputs.operational.insurance_rate, "%"),
        ]),
        ("税费与财务", [
            ("含税电价", inputs.tax_financial.tariff_with_tax, "USD/kWh"),
            ("不含税电价", inputs.tax_financial.tariff_without_tax, "USD/kWh"),
            ("增值税率", inputs.tax_financial.vat_rate, "%"),
            ("即征即退比例", inputs.tax_financial.vat_refund_rate, "%"),
            ("所得税率", inputs.tax_financial.income_tax_rate, "%"),
            ("基准折现率", inputs.tax_financial.discount_rate, "%"),
        ]),
    ]

    for section_name, items in sections:
        row += 1
        _style_cell(ws, row, 1, section_name, font=SUBHEADER_FONT, fill=SUBHEADER_FILL, align_h="left")
        for col in range(2, 5):
            _style_cell(ws, row, col, "", font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

        for name, value, unit in items:
            row += 1
            _style_cell(ws, row, 1, "", align_h="left")
            _style_cell(ws, row, 2, name, align_h="left")

            if unit == "%":
                _style_cell(ws, row, 3, value, fmt=NUM_FMT_PCT)
            elif isinstance(value, float) and abs(value) > 1000:
                _style_cell(ws, row, 3, value, fmt=NUM_FMT_USD)
            else:
                _style_cell(ws, row, 3, value)

            _style_cell(ws, row, 4, unit, align_h="left")

    _auto_width(ws)


def _create_cashflow_sheet(wb: Workbook, result: CalculationResult, inputs: WindFarmFinancialInputs):
    """逐年现金流表"""
    ws = wb.create_sheet("逐年现金流")
    ws.sheet_properties.tabColor = "548235"

    headers = [
        "年份", "营业收入", "经营成本", "折旧费", "利息支出",
        "总成本费用", "附加税金", "利润总额", "所得税", "净利润",
        "全投资税前CF", "全投资税后CF", "资本金CF", "年末贷款余额"
    ]
    _write_header_row(ws, 1, headers)

    for i, f in enumerate(result.annual_flows, start=2):
        if f.is_construction:
            label = f"建设期-{f.year + 1}"
        else:
            label = f"运营-{f.year}"

        vals = [
            label,
            f.revenue, f.total_opex, f.depreciation, f.loan_interest + f.wc_loan_interest,
            f.total_cost, f.surcharge, f.profit_before_tax, f.income_tax, f.net_profit,
            f.project_net_cf_before_tax, f.project_net_cf_after_tax, f.equity_net_cf, f.remaining_loan,
        ]
        for j, v in enumerate(vals):
            if j == 0:
                _style_cell(ws, i, j + 1, v, align_h="center")
            else:
                _style_cell(ws, i, j + 1, v, fmt=NUM_FMT_USD)

    _auto_width(ws, min_width=14)


def _create_summary_sheet(wb: Workbook, result: CalculationResult, inputs: WindFarmFinancialInputs):
    """财务指标汇总"""
    ws = wb.create_sheet("财务指标")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="财务评价指标汇总").font = Font(
        name="Arial", bold=True, size=14, color="1F4E79"
    )
    ws["A1"].alignment = Alignment(horizontal="center")

    _write_header_row(ws, 3, ["指标类别", "指标名称", "值"])

    metrics = [
        ("全投资指标", [
            ("全投资 IRR (税前)", result.project_irr_before_tax, NUM_FMT_PCT),
            ("全投资 IRR (税后)", result.project_irr_after_tax, NUM_FMT_PCT),
            ("全投资 NPV (税后)", result.project_npv_after_tax, NUM_FMT_USD),
            ("投资回收期 (税前)", result.payback_before_tax, NUM_FMT_USD2),
            ("投资回收期 (税后)", result.payback_after_tax, NUM_FMT_USD2),
        ]),
        ("资本金指标", [
            ("资本金 IRR", result.equity_irr, NUM_FMT_PCT),
            ("资本金 NPV", result.equity_npv, NUM_FMT_USD),
        ]),
        ("收益率", [
            ("度电成本 LCOE (USD/kWh)", result.lcoe, NUM_FMT_LCOE),
            ("总投资收益率 ROI", result.roi, NUM_FMT_PCT),
            ("资本金净利润率 ROE", result.roe, NUM_FMT_PCT),
        ]),
        ("汇总", [
            ("总营业收入", result.total_revenue, NUM_FMT_USD),
            ("总成本费用", result.total_cost, NUM_FMT_USD),
            ("总所得税", result.total_income_tax, NUM_FMT_USD),
        ]),
    ]

    row = 3
    for section, items in metrics:
        row += 1
        _style_cell(ws, row, 1, section, font=SUBHEADER_FONT, fill=SUBHEADER_FILL, align_h="left")
        _style_cell(ws, row, 2, "", font=SUBHEADER_FONT, fill=SUBHEADER_FILL)
        _style_cell(ws, row, 3, "", font=SUBHEADER_FONT, fill=SUBHEADER_FILL)

        for name, value, fmt in items:
            row += 1
            _style_cell(ws, row, 1, "", align_h="left")
            _style_cell(ws, row, 2, name, align_h="left")
            _style_cell(ws, row, 3, value, fmt=fmt)

    _auto_width(ws, min_width=18)


def export_to_excel(
    inputs: WindFarmFinancialInputs,
    result: CalculationResult,
    filepath: Optional[str] = None,
) -> Optional[bytes]:
    """
    导出 Excel 报告

    Args:
        inputs: 项目参数
        result: 计算结果
        filepath: 如果指定则保存到文件；否则返回 bytes 供 Streamlit 下载

    Returns:
        如果 filepath 为 None，返回 Excel 文件 bytes
    """
    wb = Workbook()

    _create_cover(wb, inputs)
    _create_params_sheet(wb, inputs)
    _create_cashflow_sheet(wb, result, inputs)
    _create_summary_sheet(wb, result, inputs)

    if filepath:
        wb.save(filepath)
        return None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
